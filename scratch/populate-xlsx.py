import openpyxl
import uuid
import datetime

# Load downloaded plantilla
wb = openpyxl.load_workbook('plantilla.xlsx')

# Inspect sheets
print("Sheets:", wb.sheetnames)

# Populate Instrumentos
ws_ins = wb['Instrumentos']
inst_uuid = str(uuid.uuid4())
print("Generated Instrument UUID:", inst_uuid)
# Header: IdInstrumento, TipoInstrumento, NombreInstrumento, TickerISIN, MonedaInstrumento, Activo
ws_ins.append([inst_uuid, 'ACCION', 'Test Instrument Inc', 'TST', 'CLP', 'SI'])

# Get Cartera ID from Carteras sheet (or use ours)
ws_car = wb['Carteras']
cartera_id = ws_car.cell(row=2, column=1).value
print("Cartera ID from template:", cartera_id)

# Populate Operaciones
ws_ops = wb['Operaciones']
# Header: IdCartera, IdInstrumento, TipoOperacionInstrumento, FechaOperacion, FechaLiquidacion, Cantidad, PrecioUnitario, MontoBruto, Comision, OtrosCargos, ImpuestoRetenido, MonedaOperacion, PrecioReferencia, PrecioEjecutado, ResultadoVsReferencia, TipoCambioOperacion, MonedaParTipoCambio, NormaTributariaRef, Observacion, RegistrarMovimientoCaja
ws_ops.append([
    cartera_id,
    inst_uuid,
    'COMPRA',
    '2026-06-25',
    '2026-06-25',
    10,
    100,
    1000,
    5,
    0,
    0,
    'CLP',
    None,
    None,
    None,
    1.0,
    'CLP',
    'TestNorma',
    'Manual test batch upload',
    'SI'
])

wb.save('plantilla_test.xlsx')
print("Saved plantilla_test.xlsx successfully.")
